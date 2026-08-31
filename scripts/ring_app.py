# ============================================================
# ring_app.py  -  standalone live GUI with configuration + calibration
# ------------------------------------------------------------
# A self-contained operator application. It does NOT modify or depend on
# the other scripts - detection, homography, offset, CSV logging, the
# homography builder and the GUI all live here.
#
# Tabs:
#   Live           annotated image + results table + log, Start / Stop
#   Configuration  folder, image ext, poll time, output CSV, XY offset,
#                  calibration (homography) file, and the FastSAM
#                  detection parameters - saved to config/app_config.json
#   Calibration    build the pixel->robot homography from point pairs:
#                  grab an image (from the watch folder or a file), the
#                  ring pixel is detected, you type the robot XY for that
#                  ring and Add point; Fit & Save writes the homography.
#
# Robot XY = homography(ring pixel) + (offset_x, offset_y).
#
# Install once:  pip install ultralytics opencv-python numpy pillow
# Run:           python scripts/ring_app.py     (or press Run in PyCharm)
# ============================================================

import os
import re
import sys
import csv
import glob
import json
import shutil
import time
import queue
import socket
import select
import fnmatch
import threading

import cv2
import numpy as np
from ultralytics import FastSAM

# tkinter / Pillow imported lazily in main() so the worker stays testable.
tk = ttk = filedialog = messagebox = simpledialog = Image = ImageTk = None

if getattr(sys, "frozen", False):
    # running as a PyInstaller .exe -> keep config/ and data/ next to the exe
    ROOT = os.path.dirname(sys.executable)
else:
    ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_FILE = os.path.join(ROOT, "config", "app_config.json")

# bundled resources (icon) live in _MEIPASS when frozen, else the project root
_RES = getattr(sys, "_MEIPASS",
               os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
ICON_FILE = os.path.join(_RES, "assets", "ring_app.ico")

DEFAULT_CONFIG = {
    "watch_folder": os.path.join(ROOT, "data", "incoming"),
    "image_ext": ".bmp",
    "name_pattern": "*",        # glob on filename; "*" = any, e.g. "202*_*.bmp"
    "poll_seconds": 0.5,
    "output_csv": os.path.join(ROOT, "data", "results.csv"),
    "latest_csv": os.path.join(ROOT, "data", "latest.csv"),  # overwritten each image
    "offset_x": 0.0,
    "offset_y": 0.0,
    "homography_file": os.path.join(ROOT, "config", "robot_map.json"),
    "intrinsics_file": os.path.join(ROOT, "config", "calibration.json"),  # optional
    "undistort": False,             # distortion-aware map (needs intrinsics_file)
    "calib_transform": "homography",  # homography | affine | similarity (sin/cos)
    "batch_multi": "largest",         # if an image has >1 ring: largest|center|skip
    # ---- detection parameters ----
    "model": "FastSAM-x.pt",
    "model_type": "auto",           # auto | fastsam | yolo | sam
    "subpixel": True,               # least-squares circle fit (more precise)
    "imgsz": 640,                   # inference size; 640 best for ~320x240 input
    "frames_avg": 1,                # average N frames of a stationary part (1=off)
    "conf": 0.20,
    "iou": 0.7,
    "min_area_frac": 0.004,
    "max_area_frac": 0.25,
    "min_circ": 0.75,
    "min_radius_frac": 0.03,
    "clahe": False,                 # normalise local contrast before detect
    "auto_retry": True,             # if 0 rings, retry once with relaxed params
    "multiscale": True,             # detect at several imgsz and merge (robust)
    "multiscale_sizes": "512,640,768",  # inference sizes used when multiscale on
    "refine_od": True,              # snap circle to the true outer metal edge
    "measure_inner": False,         # also estimate inner diameter (hole) - best effort
    "inner_sat_thresh": 70,         # metal is below this saturation; hole above
    # ---- TCP server (sends robot XY to the controller) ----
    "tcp_enabled": False,
    "tcp_host": "0.0.0.0",
    "tcp_port": 5000,
    "tcp_format": "{id},{x},{y}",   # one line per ring; \n appended
    "tcp_empty_message": "EMPTY",   # sent when no ring is found (blank = off)
    # ---- watcher housekeeping ----
    "auto_delete_minutes": 10,      # delete images older than this (0 = off)
    # ---- output ordering ----
    "sort_by": "y",                 # y | x | diameter  (order of ring ids/CSV/TCP)
    "sort_desc": False,
    # ---- read robot position over TCP (calibration) ----
    "robot_ip": "127.0.0.1",
    "robot_port": 4000,
    "robot_query": "STP",           # command sent to request position (blank = just read)
    "robot_poll": 0.3,              # seconds between position reads
    "auto_start": True,             # start watching automatically when the app opens
}


ROBOT_XY_RE = re.compile(rb"X\s*=\s*(-?\d+\.?\d*)\s*,\s*Y\s*=\s*(-?\d+\.?\d*)",
                         re.IGNORECASE)


def read_robot_xy(host, port, query, timeout=3.0):
    """Connect to the robot, optionally send `query`, read a reply and parse
    'X=xx.xx,Y=yy.yy'. Returns (x, y) or raises."""
    s = socket.create_connection((host, int(port)), timeout=timeout)
    try:
        s.settimeout(timeout)
        if query:
            s.sendall((query + "\n").encode("ascii", "replace"))
        buf = b""
        end = time.time() + timeout
        while time.time() < end:
            try:
                d = s.recv(256)
            except socket.timeout:
                break
            if not d:
                break
            buf += d
            m = ROBOT_XY_RE.search(buf)
            if m:
                return float(m.group(1)), float(m.group(2))
        m = ROBOT_XY_RE.search(buf)
        if m:
            return float(m.group(1)), float(m.group(2))
        raise ValueError("no 'X=..,Y=..' in reply: %r" % buf[:80])
    finally:
        try:
            s.close()
        except OSError:
            pass


class RobotReader(threading.Thread):
    """Keeps a live TCP connection to the robot and continuously reads its
    current X/Y position. The latest value is available via get_latest();
    ('robot_live', (x, y)) and ('robot_status', text) go to the queue."""

    def __init__(self, q, get_cfg):
        super().__init__(daemon=True)
        self.q = q
        self.get_cfg = get_cfg
        self.enabled = threading.Event()
        self.stop_flag = threading.Event()
        self._latest = None
        self._lock = threading.Lock()
        self._status = ""

    def get_latest(self):
        with self._lock:
            return self._latest

    def _set_status(self, text):
        if text != self._status:
            self._status = text
            self.q.put(("robot_status", text))

    def run(self):
        sock = None
        buf = b""
        while not self.stop_flag.is_set():
            if not self.enabled.is_set():
                if sock:
                    try:
                        sock.close()
                    except OSError:
                        pass
                    sock = None
                    self._set_status("off")
                time.sleep(0.2)
                continue
            cfg = self.get_cfg()
            host = cfg.get("robot_ip", "127.0.0.1")
            port = int(cfg.get("robot_port", 4000))
            query = cfg.get("robot_query", "")
            try:
                interval = float(cfg.get("robot_poll", 0.3) or 0.3)
            except (TypeError, ValueError):
                interval = 0.3
            try:
                if sock is None:
                    sock = socket.create_connection((host, port), timeout=3)
                    sock.settimeout(3)
                    buf = b""
                    self._set_status("connected %s:%s" % (host, port))
                if query:
                    sock.sendall((query + "\n").encode("ascii", "replace"))
                data = sock.recv(256)
                if not data:
                    raise OSError("connection closed")
                buf += data
                matches = list(ROBOT_XY_RE.finditer(buf))
                if matches:
                    m = matches[-1]
                    x, y = float(m.group(1)), float(m.group(2))
                    with self._lock:
                        self._latest = (x, y)
                    self.q.put(("robot_live", (x, y)))
                    buf = buf[m.end():][-256:]
                time.sleep(interval)
            except Exception as e:
                if sock:
                    try:
                        sock.close()
                    except OSError:
                        pass
                    sock = None
                with self._lock:
                    self._latest = None
                self._set_status("disconnected (%s)" % e)
                time.sleep(1.0)

# detection parameter keys shown on the Configuration tab (numeric)
DET_PARAMS = ["conf", "iou", "min_area_frac", "max_area_frac",
              "min_circ", "min_radius_frac"]


def load_config():
    cfg = dict(DEFAULT_CONFIG)
    try:
        cfg.update(json.load(open(CONFIG_FILE)))
    except Exception:
        pass
    return cfg


def save_config(cfg):
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    json.dump(cfg, open(CONFIG_FILE, "w"), indent=2)


def read_robot_table(path):
    """Read a calibration table (.xlsx/.xls/.csv). Finds columns for X, Y and
    image name by header keywords. Returns [(image_name, x, y), ...]."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        data = [list(r) for r in wb.active.iter_rows(values_only=True)]
    else:
        with open(path, newline="") as f:
            data = [r for r in csv.reader(f)]
    # locate the header row (has an image/name column)
    hdr_i = None
    for i, row in enumerate(data):
        cells = [str(c).lower() if c is not None else "" for c in row]
        if any("image" in c or "name" in c for c in cells):
            hdr_i = i
            break
    if hdr_i is None:
        raise ValueError("no header row with an 'image/name' column found")
    header = [str(c).lower() if c is not None else "" for c in data[hdr_i]]
    xcol = ycol = imgcol = None
    for j, h in enumerate(header):
        if imgcol is None and ("image" in h or "name" in h):
            imgcol = j
        elif xcol is None and "x" in h:
            xcol = j
        elif ycol is None and "y" in h:
            ycol = j
    if None in (xcol, ycol, imgcol):
        raise ValueError("could not find X / Y / image columns in the header")
    out = []
    for row in data[hdr_i + 1:]:
        if imgcol >= len(row) or row[imgcol] in (None, ""):
            continue
        try:
            out.append((str(row[imgcol]).strip(),
                        float(row[xcol]), float(row[ycol])))
        except (ValueError, TypeError):
            continue
    if not out:
        raise ValueError("no data rows parsed")
    return out


def load_intrinsics(path):
    """Load camera_matrix + distortion from a calibrate.py calibration.json.
    Returns (K, dist) or (None, None)."""
    try:
        d = json.load(open(path))
        K = np.array(d["camera_matrix"], np.float64)
        dist = np.array(d["distortion_coefficients"], np.float64)
        return K, dist
    except Exception:
        return None, None


def load_mapper(path):
    """Load a pixel->robot map. Returns dict {H|M, K, dist} or None. 'H' is a
    3x3 homography; 'M' is a 2x3 affine/similarity. K/dist present when the map
    is distortion-aware."""
    try:
        d = json.load(open(path))
        m = {"H": None, "M": None,
             "K": np.array(d["K"], np.float64) if d.get("K") else None,
             "dist": np.array(d["dist"], np.float64)
             if d.get("dist") is not None else None}
        if d.get("M") is not None:
            m["M"] = np.array(d["M"], np.float64)      # affine / similarity
        elif d.get("H") is not None:
            m["H"] = np.array(d["H"], np.float64)      # homography
        else:
            return None
        return m
    except Exception:
        return None


def _undistort(pts, mapper):
    pts = np.asarray(pts, np.float64).reshape(-1, 1, 2)
    if mapper.get("K") is not None:
        return cv2.undistortPoints(pts, mapper["K"], mapper["dist"], P=mapper["K"])
    return pts


def _apply_map(mapper, pts_1x1x2):
    """Apply the map (H or M) to undistorted points (N,1,2) -> (N,2)."""
    if mapper.get("M") is not None:
        M = mapper["M"]
        p = pts_1x1x2.reshape(-1, 2)
        return (p @ M[:, :2].T) + M[:, 2]
    return cv2.perspectiveTransform(pts_1x1x2, mapper["H"]).reshape(-1, 2)


def map_point(mapper, x, y):
    """Pixel (x, y) -> robot (X, Y) mm (undistorts first if intrinsics present)."""
    q = _apply_map(mapper, _undistort([(x, y)], mapper))[0]
    return float(q[0]), float(q[1])


def map_diameter_mm(mapper, x, y, r):
    """Outer diameter in mm: map opposite rim points and average."""
    mm = _apply_map(mapper,
                    _undistort([(x + r, y), (x - r, y), (x, y + r), (x, y - r)],
                               mapper))
    return 0.5 * (float(np.linalg.norm(mm[0] - mm[1])) +
                  float(np.linalg.norm(mm[2] - mm[3])))


def _estimate_transform(P, R, kind, thr, robust=True):
    """Fit pixel->robot transform. kind: homography | affine | similarity.
    Returns (key, matrix, mask) where key is 'H' (3x3) or 'M' (2x3)."""
    if kind == "homography":
        M, mask = cv2.findHomography(P, R, cv2.RANSAC if robust else 0, thr)
        return "H", M, mask
    method = cv2.RANSAC if robust else cv2.LMEDS
    if kind == "affine":
        M, mask = cv2.estimateAffine2D(P, R, method=method,
                                       ransacReprojThreshold=thr)
    else:                                   # similarity: rotation+scale+trans
        M, mask = cv2.estimateAffinePartial2D(P, R, method=method,
                                              ransacReprojThreshold=thr)
    return "M", M, mask


def _apply_transform(key, M, P):
    if key == "H":
        return cv2.perspectiveTransform(P.reshape(-1, 1, 2), M).reshape(-1, 2)
    return (P @ M[:, :2].T) + M[:, 2]


def fit_transform(points, kind="homography", ransac_mm=6.0, K=None, dist=None,
                  frame_wh=None):
    """points = [(px, py, robot_x, robot_y), ...] -> map dict + stats.
    kind: 'homography' (perspective), 'affine' (scale/shear/rot/trans) or
    'similarity' (rotation+uniform scale+translation, i.e. sin/cos - best when
    the camera views a flat plane squarely). Raises ValueError if too few
    points (homography needs >=4; affine/similarity >=3)."""
    need = 4 if kind == "homography" else 3
    if len(points) < need:
        raise ValueError("need at least %d points, have %d" % (need, len(points)))
    P = np.array([[p[0], p[1]] for p in points], np.float64)
    R = np.array([[p[2], p[3]] for p in points], np.float64)
    if K is not None and dist is not None:
        P = cv2.undistortPoints(P.reshape(-1, 1, 2), K, dist, P=K).reshape(-1, 2)
    key, M, mask = _estimate_transform(P, R, kind, ransac_mm, robust=True)
    if M is None:
        raise ValueError("%s fit failed" % kind)
    mask = mask.ravel() if mask is not None else np.ones(len(P))
    inl = [i for i in range(len(P)) if mask[i]]
    if len(inl) < need:
        inl = list(range(len(P)))
    # deterministic refit on inliers
    key, M, _ = _estimate_transform(P[inl], R[inl], kind, ransac_mm, robust=False)
    proj = _apply_transform(key, M, P[inl])
    res = np.linalg.norm(proj - R[inl], axis=1)
    loo = None
    if len(inl) >= need + 1:
        errs = []
        for h in inl:
            tr = [i for i in inl if i != h]
            k2, M2, _ = _estimate_transform(P[tr], R[tr], kind, ransac_mm,
                                            robust=False)
            q = _apply_transform(k2, M2, P[h].reshape(1, 2))[0]
            errs.append(float(np.linalg.norm(q - R[h])))
        loo = float(np.sqrt(np.mean(np.square(errs))))
    # coverage warning
    warn = None
    Pin = np.array([[points[i][0], points[i][1]] for i in inl], np.float64)
    c = Pin - Pin.mean(axis=0)
    ev = np.linalg.eigvalsh(np.cov(c.T)) if len(inl) >= 2 else np.array([1.0, 1.0])
    ratio = float(ev.min() / ev.max()) if ev.max() > 0 else 0.0
    if ratio < 0.05:
        warn = ("calibration points are nearly collinear - spread them across "
                "the whole frame for a reliable map")
    elif frame_wh:
        xr = float(Pin[:, 0].max() - Pin[:, 0].min())
        yr = float(Pin[:, 1].max() - Pin[:, 1].min())
        if xr < 0.30 * frame_wh[0] or yr < 0.30 * frame_wh[1]:
            warn = ("calibration points are clustered (cover %.0f%% x %.0f%% of "
                    "the frame) - spread them out and toward the corners"
                    % (100 * xr / frame_wh[0], 100 * yr / frame_wh[1]))
    result = {
        "type": "px_to_robot_mm",
        "transform": kind,
        "points_used": len(inl),
        "excluded": [i + 1 for i in range(len(points)) if i not in inl],
        "rms_mm": float(np.sqrt((res ** 2).mean())),
        "loo_rms_mm": loo,
        "coverage_warning": warn,
        "frame_wh": list(frame_wh) if frame_wh else None,
        "points": [{"id": i + 1, "px": p[0], "py": p[1],
                    "robot_x": p[2], "robot_y": p[3]}
                   for i, p in enumerate(points)],
    }
    if key == "H":
        result["H"] = M.tolist()
    else:
        result["M"] = M.tolist()
        result["angle_deg"] = float(np.degrees(np.arctan2(M[1, 0], M[0, 0])))
        result["scale_mm_px"] = float(np.hypot(M[0, 0], M[1, 0]))
    if K is not None and dist is not None:
        result["K"] = K.tolist()
        result["dist"] = dist.ravel().tolist()
    return result


def fit_homography(points, ransac_mm=6.0, K=None, dist=None, frame_wh=None):
    """Backward-compatible wrapper -> homography."""
    return fit_transform(points, "homography", ransac_mm, K, dist, frame_wh)


# ---------------- TCP server ----------------

class TcpServer:
    """Simple line-oriented TCP server. The robot controller connects as a
    client; each detection is broadcast to all connected clients."""

    def __init__(self, log):
        self.log = log
        self.sock = None
        self.clients = []
        self.lock = threading.Lock()
        self.running = False

    def start(self, host, port):
        self.stop()
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind((host, int(port)))
            s.listen(5)
            s.settimeout(1.0)
        except OSError as e:
            self.log("TCP: cannot listen on %s:%s (%s)" % (host, port, e))
            return False
        self.sock = s
        self.running = True
        threading.Thread(target=self._accept, daemon=True).start()
        self.log("TCP server listening on %s:%s" % (host, port))
        return True

    def _accept(self):
        while self.running:
            try:
                c, addr = self.sock.accept()
            except socket.timeout:
                continue
            except OSError:
                break
            with self.lock:
                self.clients.append((c, addr))
            self.log("TCP client connected: %s:%s" % addr)

    def broadcast(self, text):
        if not self.running or not text:
            return
        data = text.encode("ascii", "replace")
        dead = []
        with self.lock:
            for c, addr in self.clients:
                try:
                    c.sendall(data)
                except OSError:
                    dead.append((c, addr))
            for c, addr in dead:
                try:
                    c.close()
                except OSError:
                    pass
                self.clients.remove((c, addr))
        if dead:
            self.log("TCP: dropped %d disconnected client(s)" % len(dead))

    def info(self):
        """(running, [client 'ip:port' strings]) snapshot for the GUI.
        Also reaps clients that have disconnected while idle."""
        with self.lock:
            alive = []
            for c, addr in self.clients:
                try:
                    r, _, _ = select.select([c], [], [], 0)
                    if r and c.recv(1, socket.MSG_PEEK) == b"":
                        c.close()               # peer closed
                        continue
                except OSError:
                    try:
                        c.close()
                    except OSError:
                        pass
                    continue
                alive.append((c, addr))
            self.clients = alive
            return self.running, ["%s:%s" % addr for _, addr in alive]

    def stop(self):
        self.running = False
        with self.lock:
            for c, _ in self.clients:
                try:
                    c.close()
                except OSError:
                    pass
            self.clients = []
        if self.sock:
            try:
                self.sock.close()
            except OSError:
                pass
            self.sock = None


# ---------------- detection ----------------

def fit_circle_ls(pts):
    """Least-squares (Kasa) circle fit to Nx2 points -> (cx, cy, r).
    More accurate/stable than minEnclosingCircle, which is biased outward."""
    pts = np.asarray(pts, np.float64)
    x, y = pts[:, 0], pts[:, 1]
    A = np.c_[2 * x, 2 * y, np.ones(len(x))]
    b = x * x + y * y
    sol, *_ = np.linalg.lstsq(A, b, rcond=None)
    cx, cy, cc = sol
    r2 = cc + cx * cx + cy * cy
    if not np.isfinite(r2) or r2 <= 0:
        raise ValueError("bad circle fit")
    return float(cx), float(cy), float(np.sqrt(r2))


def download_asset(name, dest_dir, log=None):
    """Download a model weight (e.g. FastSAM-x.pt) into dest_dir. Returns the
    local path. Uses the ultralytics asset downloader; needs internet once."""
    base = os.path.basename(str(name))
    dest = os.path.join(dest_dir, base)
    if os.path.exists(dest):
        if log:
            log("model already present: %s" % dest)
        return dest
    from ultralytics.utils.downloads import attempt_download_asset
    p = attempt_download_asset(base)          # downloads, returns local path
    if p and os.path.exists(p):
        if os.path.abspath(p) != os.path.abspath(dest):
            os.makedirs(dest_dir, exist_ok=True)
            shutil.copy(p, dest)
        if log:
            log("model downloaded: %s" % dest)
        return dest
    raise RuntimeError("download failed for %s" % base)


class Detector:
    def __init__(self):
        self.model = None
        self.model_name = None
        self.kind = None

    @staticmethod
    def resolve_kind(name, want="auto"):
        want = (want or "auto").lower()
        if want in ("fastsam", "yolo", "sam"):
            return want
        n = os.path.basename(str(name)).lower()
        if "fastsam" in n:
            return "fastsam"
        if "sam" in n:
            return "sam"
        return "yolo"                    # yolo / custom-trained .pt

    def load(self, name, model_type="auto"):
        kind = self.resolve_kind(name, model_type)
        if self.model is not None and name == self.model_name and kind == self.kind:
            return
        if kind == "fastsam":
            from ultralytics import FastSAM
            self.model = FastSAM(name)
        elif kind == "sam":
            from ultralytics import SAM
            self.model = SAM(name)
        else:
            from ultralytics import YOLO
            self.model = YOLO(name)
        self.model_name = name
        self.kind = kind

    def _predict(self, img, conf, iou, imgsz):
        if self.kind == "fastsam":
            return self.model(img, device="cpu", retina_masks=True, imgsz=imgsz,
                              conf=conf, iou=iou, verbose=False)[0]
        return self.model(img, device="cpu", imgsz=imgsz, conf=conf, iou=iou,
                          verbose=False)[0]

    def _extract(self, res, H, W, min_r, min_af, max_af, min_circ, subpixel):
        """Turn a prediction into a list of (x, y, r) rings."""
        rings = []

        def add(x, y, r):
            if r < min_r:
                return
            # concentric detections (e.g. FastSAM giving the hole as well as the
            # washer) -> keep the LARGER one so the outer diameter always wins.
            for idx, (rx, ry, rr) in enumerate(rings):
                if (x - rx) ** 2 + (y - ry) ** 2 < (0.6 * max(r, rr)) ** 2:
                    if r > rr:
                        rings[idx] = (x, y, r)
                    return
            rings.append((x, y, r))

        masks = getattr(res, "masks", None)
        if masks is not None and masks.data is not None:
            md = masks.data.cpu().numpy()
            # largest area first, so an outer washer is added before its hole
            order = sorted(range(len(md)), key=lambda k: -(md[k] > 0.5).sum())
            for k in order:
                mask = (md[k] > 0.5).astype(np.uint8)
                a = int(mask.sum())
                if a < min_af * H * W or a > max_af * H * W:
                    continue
                cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)
                c = max(cnts, key=cv2.contourArea)
                (x, y), r = cv2.minEnclosingCircle(c)
                peri = cv2.arcLength(c, True)
                circ = 4 * np.pi * cv2.contourArea(c) / (peri * peri) if peri else 0
                if circ < min_circ:
                    continue
                if subpixel and len(c) >= 5:
                    try:
                        x, y, r = fit_circle_ls(c.reshape(-1, 2))
                    except Exception:
                        pass
                add(x, y, r)
        else:
            # detection-only model (no masks): use bounding boxes as circles
            boxes = getattr(res, "boxes", None)
            if boxes is not None:
                for b in boxes.xyxy.cpu().numpy():
                    x1, y1, x2, y2 = b
                    add((x1 + x2) / 2.0, (y1 + y2) / 2.0,
                        min(x2 - x1, y2 - y1) / 2.0)
        return sorted(rings, key=lambda t: (t[1], t[0]))

    @staticmethod
    def _merge(rings):
        """Union rings from several passes, keeping the LARGER of any two that
        are concentric (so the outer washer always wins over its hole)."""
        out = []
        for x, y, r in sorted(rings, key=lambda t: -t[2]):     # largest first
            if all((x - a) ** 2 + (y - b) ** 2 >= (0.6 * max(r, c)) ** 2
                   for a, b, c in out):
                out.append((x, y, r))
        return out

    def _sizes(self, cfg, H, W):
        """Resolve the list of inference sizes to run. FastSAM can drop a ring
        at one imgsz but catch it at another, so detecting at several sizes and
        merging is far more reliable than any single size."""
        def one(v):
            if str(v).strip().lower() == "auto":
                return int(min(1024, max(640, round(max(H, W) / 32.0) * 32)))
            return int(v)
        if bool(cfg.get("multiscale", True)):
            spec = str(cfg.get("multiscale_sizes", "512,640,768"))
            sizes = [one(s) for s in spec.split(",") if s.strip()]
            return sizes or [one(cfg.get("imgsz", 640))]
        return [one(cfg.get("imgsz", 640))]

    def find_rings(self, img, cfg):
        H, W = img.shape[:2]
        min_r = max(5, (H * W) ** 0.5 * float(cfg.get("min_radius_frac", 0.03)))
        conf = float(cfg.get("conf", 0.20))
        iou = float(cfg.get("iou", 0.7))
        min_af = float(cfg.get("min_area_frac", 0.004))
        max_af = float(cfg.get("max_area_frac", 0.25))
        min_circ = float(cfg.get("min_circ", 0.75))
        subpixel = bool(cfg.get("subpixel", True))
        # CLAHE can help on flat/low-contrast belts but can also hurt on busy
        # textured belts - off by default, fed to the model only (geometry and
        # measurement stay on the original pixels).
        det_img = apply_clahe(img) if bool(cfg.get("clahe", False)) else img

        sizes = self._sizes(cfg, H, W)
        found = []
        for imgsz in sizes:
            res = self._predict(det_img, conf, iou, imgsz)
            found += self._extract(res, H, W, min_r, min_af, max_af,
                                   min_circ, subpixel)
        rings = self._merge(found)

        # last-resort safety net: nothing at any scale -> one relaxed pass
        # (lower conf, looser circularity) before declaring the belt empty.
        if not rings and bool(cfg.get("auto_retry", True)):
            conf2 = max(0.03, conf * 0.4)
            res = self._predict(det_img, conf2, iou, max(sizes))
            rings = self._extract(res, H, W, min_r, min_af, max_af,
                                  min_circ * 0.8, subpixel)

        # snap each circle to the true outer metal edge so a thin washer reports
        # its OD, not the inner hole (measured on the ORIGINAL, un-CLAHE pixels).
        if rings and bool(cfg.get("refine_od", True)):
            gray = cv2.GaussianBlur(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                                    .astype(np.float32), (3, 3), 0)
            rings = [(x, y, refine_outer_radius(gray, x, y, r))
                     for (x, y, r) in rings]
        return sorted(rings, key=lambda t: (t[1], t[0]))


def sort_rings(rings, by="y", desc=False):
    """rings = [(x, y, r), ...]; sort by y | x | diameter."""
    idx = {"x": 0, "y": 1, "diameter": 2}.get(by, 1)
    return sorted(rings, key=lambda t: t[idx], reverse=bool(desc))


_CLAHE = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))


def apply_clahe(img):
    """Even out brightness across the frame (CLAHE on the L channel) so a ring
    in a dark corner has the same contrast as one in a bright centre. Returns a
    BGR image; used only as detection input - display/measurement stay on the
    original pixels."""
    try:
        lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
        lab[:, :, 0] = _CLAHE.apply(lab[:, :, 0])
        return cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
    except Exception:
        return img


def refine_outer_radius(gray, cx, cy, r0):
    """Snap a detected radius to the true OUTER metal edge (OD).

    FastSAM sometimes segments the inner disc / mid-wall of a thin washer, so
    the circle lands on the hole (ID) instead of the outside. We take the mean
    intensity around the ring at increasing radius and pick the OUTERMOST strong
    edge (largest |gradient|) in a band around r0 - that is where the metal meets
    the belt. Polarity-independent (works whether metal is brighter or darker).
    Returns the refined radius, or r0 unchanged when there is no clear edge."""
    H, W = gray.shape[:2]
    rs = np.arange(max(4.0, 0.55 * r0), 1.55 * r0, 0.5)
    if len(rs) < 6:
        return r0
    th = np.linspace(0, 2 * np.pi, 240, endpoint=False)
    ct, st = np.cos(th), np.sin(th)
    prof = []
    for rr in rs:
        xs = cx + rr * ct
        ys = cy + rr * st
        ok = (xs >= 0) & (xs < W - 1) & (ys >= 0) & (ys < H - 1)
        prof.append(gray[ys[ok].astype(int), xs[ok].astype(int)].mean()
                    if ok.any() else 0.0)
    prof = np.array(prof)
    g = np.abs(np.gradient(prof))
    gmax = float(g.max())
    if gmax < 2.0:                       # no real edge -> trust the detection
        return r0
    strong = np.where(g >= 0.45 * gmax)[0]
    if len(strong) == 0:
        return r0
    i = int(strong[-1])                  # outermost strong edge = outer metal edge
    if 0 < i < len(g) - 1:               # parabolic sub-sample
        a, b, c = g[i - 1], g[i], g[i + 1]
        d = a - 2 * b + c
        off = 0.5 * (a - c) / d if abs(d) > 1e-6 else 0.0
    else:
        off = 0.0
    rod = float(rs[i] + off * (rs[1] - rs[0]))
    return rod if 0.7 * r0 <= rod <= 1.55 * r0 else r0


def inner_radius(img, x, y, r, sat_thresh=70):
    """Estimate the washer hole radius (px) from the image. FastSAM masks the
    washer solid, so we use the hole showing the (saturated) background inside
    the outer circle. Returns 0 if no clear hole. Best-effort / background-
    dependent."""
    xi, yi, ri = int(round(x)), int(round(y)), int(round(r))
    H, W = img.shape[:2]
    S = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)[:, :, 1]
    Y, X = np.ogrid[:H, :W]
    disc = ((X - xi) ** 2 + (Y - yi) ** 2) <= (ri * 0.92) ** 2
    hole = (disc & (S >= sat_thresh)).astype(np.uint8)   # non-metal inside outer
    hole = cv2.morphologyEx(hole, cv2.MORPH_OPEN, np.ones((3, 3), np.uint8))
    cnts, _ = cv2.findContours(hole, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    best = 0.0
    for c in cnts:
        (hx, hy), hr = cv2.minEnclosingCircle(c)
        if (hx - xi) ** 2 + (hy - yi) ** 2 < (0.5 * ri) ** 2 \
                and 3 < hr < 0.9 * ri and hr > best:
            best = hr
    return best


def annotate(img, rings, cfg=None, mapper=None):
    """Draw rings; returns (vis, records)."""
    ox = float(cfg.get("offset_x", 0.0)) if cfg else 0.0
    oy = float(cfg.get("offset_y", 0.0)) if cfg else 0.0
    do_inner = bool(cfg.get("measure_inner")) if cfg else False
    sat = int(cfg.get("inner_sat_thresh", 70)) if cfg else 70
    vis = img.copy()
    recs = []
    for i, (x, y, r) in enumerate(rings):
        rec = {"id": i + 1, "x": round(x, 1), "y": round(y, 1),
               "diameter": round(2 * r, 1), "robot_x": "", "robot_y": "",
               "diameter_mm": "", "inner_dia": "", "inner_dia_mm": ""}
        ir = inner_radius(img, x, y, r, sat) if do_inner else 0.0
        if ir > 0:
            rec["inner_dia"] = round(2 * ir, 1)
        label = str(i + 1)
        if mapper is not None:
            rx, ry = map_point(mapper, x, y)
            rec["robot_x"] = round(rx + ox, 3)
            rec["robot_y"] = round(ry + oy, 3)
            rec["diameter_mm"] = round(map_diameter_mm(mapper, x, y, r), 3)
            if ir > 0:
                rec["inner_dia_mm"] = round(map_diameter_mm(mapper, x, y, ir), 3)
            label += " (%.1f,%.1f) D%.1f" % (rec["robot_x"], rec["robot_y"],
                                             rec["diameter_mm"])
        if ir > 0:
            cv2.circle(vis, (int(x), int(y)), int(ir), (0, 180, 255), 2)
        cv2.circle(vis, (int(x), int(y)), int(r), (0, 255, 0), 2)
        cv2.drawMarker(vis, (int(x), int(y)), (0, 0, 255), cv2.MARKER_CROSS, 18, 2)
        cv2.putText(vis, label, (int(x) - 12, int(y) - int(r) - 5),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 255, 255), 1, cv2.LINE_AA)
        recs.append(rec)
    return vis, recs


# ---------------- background worker ----------------

class Worker(threading.Thread):
    def __init__(self, q, cfg, tcp=None):
        super().__init__(daemon=True)
        self.q = q
        self.cfg = cfg
        self.tcp = tcp
        self.det = Detector()
        self.watching = threading.Event()
        self.stop_flag = threading.Event()
        self.jobs = queue.Queue()          # (kind, path): "process" | "calib"
        self._seen = {}     # name -> (size, mtime) last observed (stability)
        self._proc = {}     # name -> mtime last processed (re-fire on change)
        self._last_cleanup = 0.0
        self._avg_buf = []

    def log(self, m):
        self.q.put(("log", m))

    def run(self):
        try:
            self.log("Loading FastSAM model (first run downloads it)...")
            self.det.load(self.cfg.get("model", "FastSAM-x.pt"))
            self.log("Model ready.")
        except Exception as e:
            self.log("ERROR loading model: %s" % e)
        self.q.put(("ready", None))
        while not self.stop_flag.is_set():
            try:
                kind, path = self.jobs.get_nowait()
                if kind == "calib":
                    self._calib(path)
                elif kind == "batchcal":
                    self._batchcal(path)
                else:
                    self._process(path, allow_avg=False)   # on-demand: no averaging
                continue
            except queue.Empty:
                pass
            if self.watching.is_set():
                self._scan()
            time.sleep(max(0.05, float(self.cfg.get("poll_seconds", 0.5))))

    def start_watching(self):
        self._seen.clear()
        self._proc.clear()
        # mark pre-existing files as already processed (at their current mtime);
        # a later overwrite changes the mtime and re-fires them.
        folder = self.cfg["watch_folder"]
        try:
            for n in os.listdir(folder):
                if self._is_img(n):
                    try:
                        self._proc[n] = os.path.getmtime(os.path.join(folder, n))
                    except OSError:
                        pass
        except OSError:
            pass
        self._avg_buf = []
        self.watching.set()
        self.log("Watching %s" % folder)

    def stop_watching(self):
        self.watching.clear()
        self.log("Stopped watching.")

    def submit(self, path, kind="process"):
        self.jobs.put((kind, path))

    def _is_img(self, n):
        ext = self.cfg.get("image_ext", ".bmp").lower()
        low = n.lower()
        if not low.endswith(ext) or low.endswith("_rings.png"):
            return False
        pattern = (self.cfg.get("name_pattern") or "*").strip()
        return fnmatch.fnmatch(low, pattern.lower())

    def _scan(self):
        folder = self.cfg["watch_folder"]
        try:
            names = [n for n in os.listdir(folder) if self._is_img(n)]
        except OSError:
            return
        present = set()
        for n in sorted(names):
            present.add(n)
            path = os.path.join(folder, n)
            try:
                st = os.stat(path)
            except OSError:
                continue
            key = (st.st_size, st.st_mtime)
            if self._seen.get(n) != key:      # still being written -> wait a poll
                self._seen[n] = key
                continue
            # stable. process if new name OR the file was modified/overwritten
            if self._proc.get(n) == st.st_mtime:
                continue
            self._process(path)
            self._proc[n] = st.st_mtime
        # forget files that disappeared, so a reappearing name re-fires
        for gone in [n for n in self._seen if n not in present]:
            self._seen.pop(gone, None)
            self._proc.pop(gone, None)
        self._cleanup(folder)

    def _cleanup(self, folder):
        """Delete images (and their _rings outputs) older than N minutes."""
        try:
            mins = float(self.cfg.get("auto_delete_minutes", 0) or 0)
        except (TypeError, ValueError):
            mins = 0
        if mins <= 0:
            return
        now = time.time()
        if now - self._last_cleanup < 20:     # throttle: at most every 20 s
            return
        self._last_cleanup = now
        cutoff = now - mins * 60
        try:
            entries = os.listdir(folder)
        except OSError:
            return
        removed = 0
        for n in entries:
            if not self._is_img(n):
                continue
            path = os.path.join(folder, n)
            try:
                if os.path.getmtime(path) >= cutoff:
                    continue
                os.remove(path)
                removed += 1
                self._seen.pop(n, None)
                self._proc.pop(n, None)
                # remove matching annotated/JSON outputs if present
                stem = os.path.splitext(path)[0]
                for ext in ("_rings.png", "_rings.json"):
                    if os.path.exists(stem + ext):
                        os.remove(stem + ext)
            except OSError:
                pass
        if removed:
            self.log("housekeeping: deleted %d image(s) older than %g min"
                     % (removed, mins))

    def _detect(self, path):
        img = cv2.imread(path)
        if img is None:
            self.log("cannot read %s" % os.path.basename(path))
            return None, None
        return img, self._detect_img(img)

    def _detect_img(self, img):
        self.det.load(self.cfg.get("model", "FastSAM-x.pt"),
                      self.cfg.get("model_type", "auto"))
        return self.det.find_rings(img, self.cfg)

    def _process(self, path, allow_avg=True):
        name = os.path.basename(path)
        try:
            img = cv2.imread(path)
            if img is None:
                self.log("cannot read %s" % name)
                return
            # frame averaging: pool N consecutive frames of a stationary part
            n = int(self.cfg.get("frames_avg", 1) or 1)
            if allow_avg and n > 1:
                if self._avg_buf and self._avg_buf[0].shape != img.shape:
                    self._avg_buf = []          # size changed -> reset
                self._avg_buf.append(img.astype(np.float32))
                if len(self._avg_buf) < n:
                    self.log("%s -> buffered %d/%d for averaging"
                             % (name, len(self._avg_buf), n))
                    return
                img = (np.mean(self._avg_buf, axis=0)).astype(np.uint8)
                self._avg_buf = []
                name = "%s (avg of %d)" % (name, n)
            t0 = time.time()
            rings = self._detect_img(img)
            detect_ms = (time.time() - t0) * 1000.0
            rings = sort_rings(rings, self.cfg.get("sort_by", "y"),
                               self.cfg.get("sort_desc", False))
            mapper = load_mapper(self.cfg.get("homography_file", ""))
            vis, recs = annotate(img, rings, self.cfg, mapper)  # mm/robot XY + dia mm
            total_ms = (time.time() - t0) * 1000.0
            self._write_csv(name, recs)
            self._write_latest_csv(name, recs)
            self._tcp_send(name, recs)
            self.q.put(("result", {"name": name, "image": vis, "rings": recs,
                                   "has_map": mapper is not None,
                                   "empty": len(recs) == 0,
                                   "detect_ms": detect_ms,
                                   "total_ms": total_ms}))
            timing = "  [detect %.0f ms, total %.0f ms]" % (detect_ms, total_ms)
            if not recs:
                self.log("%s -> CONVEYOR EMPTY (no ring)%s" % (name, timing))
            else:
                self.log("%s -> %d ring(s)%s%s" % (name, len(recs),
                         "" if mapper is not None else "  (no homography loaded!)",
                         timing))
        except Exception as e:
            self.log("ERROR on %s: %s" % (name, e))

    def _calib(self, path):
        name = os.path.basename(path)
        try:
            img, rings = self._detect(path)
            if img is None:
                return
            vis, _ = annotate(img, rings, self.cfg, None)
            self.q.put(("calib_result", {"name": name, "image": vis,
                                         "rings": [(round(x, 1), round(y, 1),
                                                    round(r, 1))
                                                   for x, y, r in rings]}))
            self.log("calib image %s -> %d ring(s)" % (name, len(rings)))
        except Exception as e:
            self.log("calib ERROR on %s: %s" % (name, e))

    def _find_image(self, folder, name):
        """Resolve an image name from the table (with or without extension)."""
        if os.path.exists(os.path.join(folder, name)):
            return os.path.join(folder, name)
        for ext in (".bmp", ".png", ".jpg", ".jpeg", ".tif", ".tiff"):
            p = os.path.join(folder, name + ext)
            if os.path.exists(p):
                return p
        return None

    def _batchcal(self, folder):
        """Batch calibration: read the table (xlsx/csv) in `folder`, detect the
        ring in each named image, fit the transform and save the map."""
        try:
            tables = []
            for ext in ("*.xlsx", "*.xls", "*.csv"):
                tables += glob.glob(os.path.join(folder, ext))
            if not tables:
                self.q.put(("batchcal_result",
                            {"error": "no .xlsx/.csv table found in the folder"}))
                return
            rows = read_robot_table(tables[0])
            self.log("batch calibration: %d rows from %s"
                     % (len(rows), os.path.basename(tables[0])))
            strat = self.cfg.get("batch_multi", "largest")
            pts, missing, multi = [], [], []
            for iname, rx, ry in rows:
                path = self._find_image(folder, iname)
                if path is None:
                    missing.append("%s (no image)" % iname)
                    continue
                img = cv2.imread(path)
                if img is None:
                    missing.append("%s (unreadable)" % iname)
                    continue
                rings = self._detect_img(img)
                if not rings:
                    missing.append("%s (no ring)" % iname)
                    continue
                fw = (img.shape[1], img.shape[0])
                if len(rings) > 1:
                    multi.append("%s (%d)" % (iname, len(rings)))
                    if strat == "skip":
                        missing.append("%s (%d rings - skipped)"
                                       % (iname, len(rings)))
                        continue
                    if strat == "center":     # ring nearest the image centre
                        cx, cy = img.shape[1] / 2.0, img.shape[0] / 2.0
                        x, y, _ = min(rings, key=lambda t: (t[0] - cx) ** 2
                                      + (t[1] - cy) ** 2)
                    else:                     # largest
                        x, y, _ = max(rings, key=lambda t: t[2])
                else:
                    x, y, _ = rings[0]
                pts.append((x, y, rx, ry))
            need = 4 if self.cfg.get("calib_transform", "homography") == \
                "homography" else 3
            if len(pts) < need:
                self.q.put(("batchcal_result",
                            {"error": "only %d usable points (need >= %d)"
                             % (len(pts), need), "missing": missing}))
                return
            K = dist = None
            if self.cfg.get("undistort"):
                K, dist = load_intrinsics(self.cfg.get("intrinsics_file", ""))
            kind = self.cfg.get("calib_transform", "homography")
            res = fit_transform(pts, kind, K=K, dist=dist, frame_wh=fw)
            path = self.cfg.get("homography_file", "")
            os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
            json.dump(res, open(path, "w"), indent=2)
            res["_saved_to"] = path
            res["_missing"] = missing
            res["_multi"] = multi
            res["_multi_strategy"] = strat
            res["_table"] = os.path.basename(tables[0])
            self.q.put(("batchcal_result", res))
            self.log("batch calibration done: %d points, RMS %.2f mm, saved %s"
                     % (res["points_used"], res["rms_mm"], os.path.basename(path)))
        except Exception as e:
            self.q.put(("batchcal_result", {"error": str(e)}))

    def _tcp_send(self, image, rings):
        if not self.tcp:
            return
        if not rings:                    # conveyor empty -> send the empty signal
            msg = self.cfg.get("tcp_empty_message", "EMPTY")
            if msg:
                self.tcp.broadcast(msg + "\n")
            return
        fmt = self.cfg.get("tcp_format", "{id},{x},{y}")
        lines = []
        for r in rings:
            if r["robot_x"] == "":       # only send rings with a robot XY
                continue
            try:
                lines.append(fmt.format(id=r["id"], x=r["robot_x"],
                                        y=r["robot_y"], dia=r["diameter"],
                                        dia_mm=r.get("diameter_mm", ""),
                                        image=image))
            except Exception:
                lines.append("%s,%s,%s" % (r["id"], r["robot_x"], r["robot_y"]))
        if lines:
            self.tcp.broadcast("\n".join(lines) + "\n")

    def _write_csv(self, image, rings):
        path = self.cfg.get("output_csv", "")
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
            new = not os.path.exists(path)
            with open(path, "a", newline="") as f:
                w = csv.writer(f)
                if new:
                    w.writerow(["timestamp", "image", "ring_id", "x_px",
                                "y_px", "diameter_px", "robot_x", "robot_y",
                                "diameter_mm", "inner_dia_px", "inner_dia_mm"])
                ts = time.strftime("%Y-%m-%d %H:%M:%S")
                for r in rings:
                    w.writerow([ts, image, r["id"], r["x"], r["y"],
                                r["diameter"], r["robot_x"], r["robot_y"],
                                r.get("diameter_mm", ""), r.get("inner_dia", ""),
                                r.get("inner_dia_mm", "")])
        except Exception as e:
            self.log("CSV write failed: %s" % e)

    def _write_latest_csv(self, image, rings):
        """Overwrite a single-image CSV holding only the current readings."""
        path = self.cfg.get("latest_csv", "")
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
            with open(path, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["timestamp", "image", "ring_id", "x_px", "y_px",
                            "diameter_px", "robot_x", "robot_y", "diameter_mm",
                            "inner_dia_px", "inner_dia_mm"])
                ts = time.strftime("%Y-%m-%d %H:%M:%S")
                for r in rings:
                    w.writerow([ts, image, r["id"], r["x"], r["y"],
                                r["diameter"], r["robot_x"], r["robot_y"],
                                r.get("diameter_mm", ""), r.get("inner_dia", ""),
                                r.get("inner_dia_mm", "")])
        except Exception as e:
            self.log("latest CSV write failed: %s" % e)


# ---------------- GUI ----------------

class App:
    def __init__(self, root):
        self.root = root
        root.title("Ring Finder - live / configuration / calibration")
        try:
            root.iconbitmap(ICON_FILE)
        except Exception:
            pass
        root.geometry("1040x720")
        self.cfg = load_config()
        self.q = queue.Queue()
        self.tcp = TcpServer(lambda m: self.q.put(("log", m)))
        self.worker = Worker(self.q, self.cfg, self.tcp)
        self.worker.start()
        self.robot_reader = RobotReader(self.q, lambda: self.cfg)
        self.robot_reader.start()
        if self.cfg.get("tcp_enabled"):
            self.tcp.start(self.cfg.get("tcp_host", "0.0.0.0"),
                           self.cfg.get("tcp_port", 5000))
        self._imgtk = None
        self._calib_imgtk = None
        self._calib_rings = []        # detected rings in current calib image
        self.points = []              # collected (px,py,rx,ry)

        nb = ttk.Notebook(root)
        nb.pack(fill=tk.BOTH, expand=True)
        self.live = ttk.Frame(nb)
        self.conf = ttk.Frame(nb)
        self.calib = ttk.Frame(nb)
        nb.add(self.live, text="  Live  ")
        nb.add(self.conf, text="  Configuration  ")
        nb.add(self.calib, text="  Calibration  ")
        self._build_live(self.live)
        self._build_config(self.conf)
        self._build_calib(self.calib)

        self.root.after(150, self._pump)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # ---- Live tab ----
    def _build_live(self, p):
        bar = ttk.Frame(p, padding=6)
        bar.pack(side=tk.TOP, fill=tk.X)
        self.start_btn = ttk.Button(bar, text="Start", command=self.start,
                                    state=tk.DISABLED)
        self.start_btn.pack(side=tk.LEFT)
        self.stop_btn = ttk.Button(bar, text="Stop", command=self.stop,
                                   state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=4)
        ttk.Button(bar, text="Open image...", command=self.open_image).pack(
            side=tk.LEFT, padx=(10, 0))
        self.status = ttk.Label(bar, text="Loading model...",
                                foreground="#b26b00")
        self.status.pack(side=tk.RIGHT)
        self.tcp_status = ttk.Label(bar, text="TCP: disabled",
                                    foreground="#777")
        self.tcp_status.pack(side=tk.RIGHT, padx=12)

        mid = ttk.Frame(p)
        mid.pack(fill=tk.BOTH, expand=True)
        left = ttk.LabelFrame(mid, text="Latest image", padding=6)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.banner = tk.Label(left, text="waiting...", font=("TkDefaultFont", 13, "bold"),
                               fg="white", bg="#666")
        self.banner.pack(fill=tk.X)
        zc = ttk.Frame(left)
        zc.pack(fill=tk.X, pady=2)
        ttk.Button(zc, text="-", width=3, command=lambda: self._zoom(-0.25)).pack(side=tk.LEFT)
        ttk.Button(zc, text="+", width=3, command=lambda: self._zoom(0.25)).pack(side=tk.LEFT, padx=2)
        ttk.Button(zc, text="Fit", command=self._zoom_fit).pack(side=tk.LEFT)
        ttk.Button(zc, text="100%", command=lambda: self._zoom_set(1.0)).pack(side=tk.LEFT, padx=2)
        self.zoom_lbl = ttk.Label(zc, text="200%")
        self.zoom_lbl.pack(side=tk.LEFT, padx=6)
        self.timing_lbl = ttk.Label(zc, text="", foreground="#555")
        self.timing_lbl.pack(side=tk.RIGHT)
        cvf = ttk.Frame(left)
        cvf.pack(fill=tk.BOTH, expand=True)
        self.img_canvas = tk.Canvas(cvf, background="#222", highlightthickness=0)
        vsb = ttk.Scrollbar(cvf, orient=tk.VERTICAL, command=self.img_canvas.yview)
        hsb = ttk.Scrollbar(cvf, orient=tk.HORIZONTAL, command=self.img_canvas.xview)
        self.img_canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.img_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._img_item = self.img_canvas.create_image(0, 0, anchor=tk.NW)
        self.zoom = 2.0
        self._last_bgr = None
        # mouse wheel zoom (Windows/Mac: <MouseWheel>; Linux: Button-4/5)
        self.img_canvas.bind("<MouseWheel>",
                             lambda e: self._zoom(0.25 if e.delta > 0 else -0.25))
        self.img_canvas.bind("<Button-4>", lambda e: self._zoom(0.25))
        self.img_canvas.bind("<Button-5>", lambda e: self._zoom(-0.25))
        right = ttk.LabelFrame(mid, text="Rings (pixel + robot mm)", padding=6)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=6, pady=6)
        cols = ("id", "x_px", "y_px", "dia_px", "robot_x", "robot_y", "dia_mm",
                "in_dia_px", "in_dia_mm")
        self.tree = ttk.Treeview(right, columns=cols, show="headings", height=12)
        for c, w in zip(cols, (28, 50, 50, 52, 78, 78, 64, 60, 62)):
            self.tree.heading(c, text=c,
                              command=lambda cc=c: self._sort_tree(cc))
            self.tree.column(c, width=w, anchor=tk.CENTER)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self._tree_sort = (None, False)   # (column, descending)

        logf = ttk.LabelFrame(p, text="Log", padding=4)
        logf.pack(side=tk.BOTTOM, fill=tk.X, padx=6, pady=(0, 6))
        self.logbox = tk.Text(logf, height=6, state=tk.DISABLED,
                              background="#111", foreground="#ddd")
        self.logbox.pack(fill=tk.X)

    # ---- Configuration tab ----
    def _build_config(self, page):
        self.vars = {}

        # fixed button bar at the BOTTOM (packed first so it is always visible)
        btns = ttk.Frame(page, padding=(10, 6))
        btns.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Separator(page, orient=tk.HORIZONTAL).pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Button(btns, text="Save settings", command=self.save).pack(side=tk.LEFT)
        ttk.Button(btns, text="Save & Restart",
                   command=self.save_and_restart).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="Restart app",
                   command=self.restart_app).pack(side=tk.LEFT)
        ttk.Button(btns, text="Reload calibration",
                   command=self._refresh_map_info).pack(side=tk.LEFT, padx=8)
        ttk.Button(btns, text="Export config...",
                   command=self.export_config).pack(side=tk.LEFT)
        ttk.Button(btns, text="Import config...",
                   command=self.import_config).pack(side=tk.LEFT, padx=6)

        # sub-tabs fill the rest
        sub = ttk.Notebook(page)
        sub.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        t_general = ttk.Frame(sub, padding=10)
        t_det = ttk.Frame(sub, padding=10)
        t_out = ttk.Frame(sub, padding=10)
        t_tcp = ttk.Frame(sub, padding=10)
        t_robot = ttk.Frame(sub, padding=10)
        sub.add(t_general, text="  General  ")
        sub.add(t_det, text="  Detection  ")
        sub.add(t_out, text="  Output  ")
        sub.add(t_tcp, text="  TCP  ")
        sub.add(t_robot, text="  Robot  ")

        # --- General ---
        general = [
            ("watch_folder", "Watch folder", "dir"),
            ("image_ext", "Image extension", "text"),
            ("name_pattern", "Filename pattern (glob)", "text"),
            ("poll_seconds", "Poll time (seconds)", "text"),
            ("output_csv", "Output CSV file (append log)", "savefile"),
            ("latest_csv", "Latest CSV file (overwritten)", "savefile"),
            ("offset_x", "Offset X (mm)", "text"),
            ("offset_y", "Offset Y (mm)", "text"),
            ("auto_delete_minutes", "Auto-delete images older than (min)", "text"),
            ("homography_file", "Calibration (homography) file", "openfile"),
            ("intrinsics_file", "Camera intrinsics file (optional)", "openfile"),
        ]
        for i, (key, label, kind) in enumerate(general):
            self._config_row(t_general, i, key, label, kind)
        self.auto_start_var = tk.BooleanVar(value=bool(self.cfg.get("auto_start", True)))
        ttk.Checkbutton(t_general, text="Start watching automatically on open",
                        variable=self.auto_start_var).grid(
            row=len(general), column=1, sticky=tk.W, pady=(6, 0))
        self.map_info = ttk.Label(t_general, text="", foreground="#555")
        self.map_info.grid(row=len(general) + 1, column=1, sticky=tk.W, pady=(6, 0))
        ttk.Label(t_general,
                  text="Robot XY = homography(pixel) + (offset X, offset Y).",
                  foreground="#777").grid(row=len(general) + 2, column=1,
                                          sticky=tk.W, pady=(2, 0))
        self._refresh_map_info()

        # --- Detection ---
        self._config_row(t_det, 0, "model", "MODEL (.pt)", "openfile")
        ttk.Button(t_det, text="Download",
                   command=self.download_model).grid(row=0, column=3, padx=4)
        ttk.Label(t_det, text="Model type", width=26).grid(row=1, column=0,
                                                           sticky=tk.W, pady=4)
        self.model_type_var = tk.StringVar(value=self.cfg.get("model_type", "auto"))
        ttk.Combobox(t_det, textvariable=self.model_type_var, width=14,
                     state="readonly",
                     values=["auto", "fastsam", "yolo", "sam"]).grid(
            row=1, column=1, sticky=tk.W)
        for i, key in enumerate(DET_PARAMS):
            self._config_row(t_det, i + 2, key, key.upper(), "text")
        r0 = len(DET_PARAMS) + 2
        self.subpixel_var = tk.BooleanVar(value=bool(self.cfg.get("subpixel", True)))
        ttk.Checkbutton(t_det, text="Sub-pixel circle fit (more precise)",
                        variable=self.subpixel_var).grid(
            row=r0, column=1, sticky=tk.W, pady=(6, 0))
        r0 += 1
        self.multiscale_var = tk.BooleanVar(
            value=bool(self.cfg.get("multiscale", True)))
        ttk.Checkbutton(t_det,
                        text="Multi-scale detect - detect at several sizes and "
                             "merge (fixes rings missed in some positions)",
                        variable=self.multiscale_var).grid(
            row=r0, column=1, sticky=tk.W, pady=(6, 0))
        r0 += 1
        self._config_row(t_det, r0, "multiscale_sizes",
                         "Multi-scale sizes (comma)", "text")
        r0 += 1
        self.refine_od_var = tk.BooleanVar(
            value=bool(self.cfg.get("refine_od", True)))
        ttk.Checkbutton(t_det,
                        text="Snap to outer edge (OD) - fixes thin rings "
                             "measured as the inner hole (ID)",
                        variable=self.refine_od_var).grid(
            row=r0, column=1, sticky=tk.W, pady=(2, 0))
        r0 += 1
        self.auto_retry_var = tk.BooleanVar(
            value=bool(self.cfg.get("auto_retry", True)))
        ttk.Checkbutton(t_det,
                        text="Auto-retry if nothing found (lower conf) before "
                             "calling belt empty",
                        variable=self.auto_retry_var).grid(
            row=r0, column=1, sticky=tk.W, pady=(2, 0))
        r0 += 1
        self.clahe_var = tk.BooleanVar(value=bool(self.cfg.get("clahe", False)))
        ttk.Checkbutton(t_det,
                        text="Even out lighting before detect (CLAHE) - helps "
                             "flat belts, can hurt busy/textured belts",
                        variable=self.clahe_var).grid(
            row=r0, column=1, sticky=tk.W, pady=(2, 0))
        r0 += 1
        self._config_row(t_det, r0, "imgsz", "Inference size (imgsz)", "text")
        r0 += 1
        self._config_row(t_det, r0, "frames_avg", "Frame averaging (N frames)", "text")
        r0 += 1
        ttk.Label(t_det, text="imgsz 640 suits ~320x240 input; use a number or "
                             "'auto' (matches the camera resolution, capped 1024). "
                             "Frame averaging pools N frames of a STATIONARY part "
                             "to cut noise (1 = off; use only when the part is "
                             "still for N frames). Recalibrate after changing "
                             "camera/resolution.",
                  foreground="#777", wraplength=520, justify=tk.LEFT).grid(
            row=r0, column=1, sticky=tk.W)
        r0 += 1
        self.measure_inner_var = tk.BooleanVar(
            value=bool(self.cfg.get("measure_inner")))
        ttk.Checkbutton(t_det, text="Also measure inner diameter (hole)",
                        variable=self.measure_inner_var).grid(
            row=r0, column=1, sticky=tk.W, pady=(6, 0))
        self._config_row(t_det, r0 + 1, "inner_sat_thresh",
                         "Inner hole saturation thresh", "text")
        ttk.Label(t_det, text="MODEL change reloads the model. Inner diameter is "
                             "estimated from the image (best-effort, background-"
                             "dependent); outer diameter is the reliable one.",
                  foreground="#777", wraplength=520, justify=tk.LEFT).grid(
            row=r0 + 2, column=1, sticky=tk.W, pady=(4, 0))

        # --- Output ---
        ttk.Label(t_out, text="Sort rings by", width=26).grid(row=0, column=0,
                                                             sticky=tk.W)
        self.sort_by_var = tk.StringVar(value=self.cfg.get("sort_by", "y"))
        ttk.Combobox(t_out, textvariable=self.sort_by_var, width=14,
                     state="readonly",
                     values=["y", "x", "diameter"]).grid(row=0, column=1,
                                                         sticky=tk.W)
        self.sort_desc_var = tk.BooleanVar(value=bool(self.cfg.get("sort_desc")))
        ttk.Checkbutton(t_out, text="descending",
                        variable=self.sort_desc_var).grid(row=0, column=2, padx=10)
        ttk.Label(t_out, text="Sets the order of ring ids in the table, CSV and "
                             "TCP output. (Click a Live column header to re-sort "
                             "the view.)", foreground="#777").grid(
            row=1, column=1, columnspan=2, sticky=tk.W, pady=(4, 0))

        # --- TCP ---
        self.tcp_enabled_var = tk.BooleanVar(value=bool(self.cfg.get("tcp_enabled")))
        ttk.Checkbutton(t_tcp, text="Enable TCP server",
                        variable=self.tcp_enabled_var).grid(row=0, column=1,
                                                            sticky=tk.W, pady=4)
        self._config_row(t_tcp, 1, "tcp_host", "Host / bind address", "text")
        self._config_row(t_tcp, 2, "tcp_port", "Port", "text")
        self._config_row(t_tcp, 3, "tcp_format", "Line format", "text")
        self._config_row(t_tcp, 4, "tcp_empty_message", "Empty message (no ring)",
                         "text")
        ttk.Label(t_tcp, text="Placeholders: {id} {x} {y} {dia} {image}. One line "
                             "per ring. Empty message sent when nothing is found "
                             "(blank = send nothing).",
                  foreground="#777").grid(row=5, column=1, sticky=tk.W)
        self.tcp_info = ttk.Label(t_tcp, text="", foreground="#555")
        self.tcp_info.grid(row=6, column=1, sticky=tk.W, pady=(2, 0))

        # --- Robot ---
        self._config_row(t_robot, 0, "robot_ip", "Robot IP", "text")
        self._config_row(t_robot, 1, "robot_port", "Robot port", "text")
        self._config_row(t_robot, 2, "robot_query", "Request command", "text")
        self._config_row(t_robot, 3, "robot_poll", "Read interval (seconds)", "text")
        ttk.Label(t_robot, text="App connects and continuously reads the position, "
                              "parsing a reply like  X=12.34,Y=56.78 . In the "
                              "Calibration tab, 'Update' copies the live value "
                              "into the entry fields.",
                  foreground="#777").grid(row=4, column=1, sticky=tk.W)

    def _config_row(self, parent, i, key, label, kind):
        ttk.Label(parent, text=label, width=26).grid(row=i, column=0,
                                                     sticky=tk.W, pady=4)
        var = tk.StringVar(value=str(self.cfg.get(key, "")))
        self.vars[key] = var
        ttk.Entry(parent, textvariable=var, width=54).grid(row=i, column=1, pady=4)
        if kind == "dir":
            ttk.Button(parent, text="Browse", command=lambda k=key:
                       self._pick_dir(k)).grid(row=i, column=2, padx=6)
        elif kind == "openfile":
            ttk.Button(parent, text="Browse", command=lambda k=key:
                       self._pick_file(k, save=False)).grid(row=i, column=2, padx=6)
        elif kind == "savefile":
            ttk.Button(parent, text="Browse", command=lambda k=key:
                       self._pick_file(k, save=True)).grid(row=i, column=2, padx=6)

    # ---- Calibration tab ----
    def _build_calib(self, p):
        top = ttk.Frame(p, padding=8)
        top.pack(fill=tk.X)
        ttk.Button(top, text="Grab latest from folder",
                   command=self.calib_grab).pack(side=tk.LEFT)
        ttk.Button(top, text="Load image...",
                   command=self.calib_load).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Batch (folder + Excel)",
                   command=self.calib_batch).pack(side=tk.LEFT, padx=(14, 0))
        self.batch_multi_var = tk.StringVar(
            value=self.cfg.get("batch_multi", "largest"))
        ttk.Label(top, text="if many rings:").pack(side=tk.LEFT, padx=(6, 2))
        ttk.Combobox(top, textvariable=self.batch_multi_var, width=8,
                     state="readonly",
                     values=["largest", "center", "skip"]).pack(side=tk.LEFT)
        ttk.Label(top, text="Detected ring:").pack(side=tk.LEFT, padx=(14, 2))
        self.calib_ring_sel = tk.Spinbox(top, from_=1, to=1, width=4)
        self.calib_ring_sel.pack(side=tk.LEFT)
        self.calib_pix = ttk.Label(top, text="pixel: -", foreground="#333")
        self.calib_pix.pack(side=tk.LEFT, padx=10)

        mid = ttk.Frame(p)
        mid.pack(fill=tk.BOTH, expand=True)
        left = ttk.LabelFrame(mid, text="Calibration image", padding=6)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.calib_canvas = tk.Label(left, background="#222")
        self.calib_canvas.pack(fill=tk.BOTH, expand=True)

        right = ttk.LabelFrame(mid, text="Collected points", padding=6)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=6, pady=6)

        # profile + distortion-aware
        prof = ttk.Frame(right)
        prof.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(prof, text="Profile").pack(side=tk.LEFT)
        self.profile_var = tk.StringVar()
        self.profile_combo = ttk.Combobox(prof, textvariable=self.profile_var,
                                           width=20, state="readonly")
        self.profile_combo.pack(side=tk.LEFT, padx=4)
        self.profile_combo.bind("<<ComboboxSelected>>",
                                lambda e: self._profile_selected())
        ttk.Button(prof, text="Save as...",
                   command=self.calib_save_as_profile).pack(side=tk.LEFT)
        self.undistort_var = tk.BooleanVar(value=bool(self.cfg.get("undistort")))
        ttk.Checkbutton(prof, text="distortion-aware",
                        variable=self.undistort_var).pack(side=tk.RIGHT)
        self.transform_var = tk.StringVar(
            value=self.cfg.get("calib_transform", "homography"))
        ttk.Combobox(prof, textvariable=self.transform_var, width=12,
                     state="readonly",
                     values=["homography", "affine", "similarity"]).pack(
            side=tk.RIGHT, padx=6)
        ttk.Label(prof, text="Model").pack(side=tk.RIGHT)

        # live robot position over TCP
        livef = ttk.Frame(right)
        livef.pack(fill=tk.X, pady=(0, 4))
        self.robot_live_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(livef, text="Read robot position (TCP)",
                        variable=self.robot_live_var,
                        command=self._toggle_robot_read).pack(side=tk.LEFT)
        self.robot_live_lbl = ttk.Label(livef, text="live: --",
                                        foreground="#777")
        self.robot_live_lbl.pack(side=tk.LEFT, padx=8)
        ttk.Button(livef, text="Update ↓",
                   command=self.calib_update_from_live).pack(side=tk.RIGHT)

        # robot XY entry (manual or filled by Update)
        entry = ttk.Frame(right)
        entry.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(entry, text="Robot X").grid(row=0, column=0, padx=2)
        self.rx_var = tk.StringVar()
        ttk.Entry(entry, textvariable=self.rx_var, width=10).grid(row=0, column=1)
        ttk.Label(entry, text="Robot Y").grid(row=0, column=2, padx=2)
        self.ry_var = tk.StringVar()
        ttk.Entry(entry, textvariable=self.ry_var, width=10).grid(row=0, column=3)
        ttk.Button(entry, text="Add point",
                   command=self.calib_add).grid(row=0, column=4, padx=6)

        cols = ("#", "px", "py", "robot_x", "robot_y")
        self.ptree = ttk.Treeview(right, columns=cols, show="headings", height=10)
        for c, w in zip(cols, (30, 60, 60, 80, 80)):
            self.ptree.heading(c, text=c)
            self.ptree.column(c, width=w, anchor=tk.CENTER)
        self.ptree.pack(fill=tk.BOTH, expand=True)

        row2 = ttk.Frame(right)
        row2.pack(fill=tk.X, pady=6)
        ttk.Button(row2, text="Remove selected",
                   command=self.calib_remove).pack(side=tk.LEFT)
        ttk.Button(row2, text="Clear all",
                   command=self.calib_clear).pack(side=tk.LEFT, padx=6)
        ttk.Button(row2, text="Load points from map",
                   command=self.calib_load_points).pack(side=tk.LEFT)
        self.calib_count = ttk.Label(row2, text="Points: 0  (need >= 4)",
                                     foreground="#b26b00")
        self.calib_count.pack(side=tk.RIGHT)

        row3 = ttk.Frame(right)
        row3.pack(fill=tk.X)
        ttk.Button(row3, text="Fit & Save homography",
                   command=self.calib_fit).pack(side=tk.LEFT)
        ttk.Button(row3, text="Coverage map",
                   command=self.calib_coverage).pack(side=tk.LEFT, padx=6)
        self.cov_overlay_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row3, text="overlay on image",
                        variable=self.cov_overlay_var).pack(side=tk.LEFT)
        self.calib_result = ttk.Label(right, text="", foreground="#333",
                                      wraplength=360, justify=tk.LEFT)
        self.calib_result.pack(fill=tk.X, pady=(6, 0))

        # verify: predicted (from selected ring) vs live robot position
        vrow = ttk.Frame(right)
        vrow.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(vrow, text="Verify (predicted vs live)",
                   command=self.calib_verify).pack(side=tk.LEFT)
        self.verify_lbl = ttk.Label(right, text="", foreground="#333",
                                    wraplength=360, justify=tk.LEFT)
        self.verify_lbl.pack(fill=tk.X, pady=(4, 0))

        self._refresh_profiles()

    # ---- config helpers ----
    def _pick_dir(self, key):
        d = filedialog.askdirectory(initialdir=self.vars[key].get() or ROOT)
        if d:
            self.vars[key].set(d)

    def _pick_file(self, key, save):
        if save:
            f = filedialog.asksaveasfilename(initialdir=ROOT,
                    defaultextension=".csv",
                    filetypes=[("CSV", "*.csv"), ("all", "*.*")])
        else:
            f = filedialog.askopenfilename(initialdir=ROOT,
                    filetypes=[("JSON", "*.json"), ("all", "*.*")])
        if f:
            self.vars[key].set(f)
            if key == "homography_file":
                self._refresh_map_info()

    def _refresh_map_info(self):
        path = self.vars["homography_file"].get()
        try:
            d = json.load(open(path))
            has = d.get("H") is not None or d.get("M") is not None
            info = ("map OK [%s]" % d.get("transform", "homography")) if has \
                else "no transform in file"
            if d.get("rms_mm") is not None:
                info += "  |  fit RMS %.2f mm" % d["rms_mm"]
            if d.get("loo_rms_mm") is not None:
                info += "  |  LOO %.2f mm" % d["loo_rms_mm"]
            self.map_info.config(text=info, foreground="#1a7f37")
        except Exception:
            self.map_info.config(text="calibration file not found / invalid",
                                 foreground="#cf222e")

    def _read_fields(self):
        """Read GUI fields into self.cfg. Returns True on success."""
        try:
            self.cfg["watch_folder"] = self.vars["watch_folder"].get()
            self.cfg["image_ext"] = self.vars["image_ext"].get() or ".bmp"
            self.cfg["name_pattern"] = self.vars["name_pattern"].get() or "*"
            self.cfg["poll_seconds"] = float(self.vars["poll_seconds"].get())
            self.cfg["output_csv"] = self.vars["output_csv"].get()
            self.cfg["latest_csv"] = self.vars["latest_csv"].get()
            self.cfg["offset_x"] = float(self.vars["offset_x"].get())
            self.cfg["offset_y"] = float(self.vars["offset_y"].get())
            self.cfg["auto_delete_minutes"] = float(
                self.vars["auto_delete_minutes"].get())
            self.cfg["homography_file"] = self.vars["homography_file"].get()
            self.cfg["intrinsics_file"] = self.vars["intrinsics_file"].get()
            self.cfg["undistort"] = bool(self.undistort_var.get())
            self.cfg["model"] = self.vars["model"].get() or "FastSAM-x.pt"
            for k in DET_PARAMS:
                self.cfg[k] = float(self.vars[k].get())
            self.cfg["measure_inner"] = bool(self.measure_inner_var.get())
            self.cfg["inner_sat_thresh"] = int(self.vars["inner_sat_thresh"].get())
            self.cfg["model_type"] = self.model_type_var.get() or "auto"
            self.cfg["subpixel"] = bool(self.subpixel_var.get())
            self.cfg["clahe"] = bool(self.clahe_var.get())
            self.cfg["auto_retry"] = bool(self.auto_retry_var.get())
            self.cfg["multiscale"] = bool(self.multiscale_var.get())
            self.cfg["multiscale_sizes"] = (
                self.vars["multiscale_sizes"].get().strip() or "512,640,768")
            self.cfg["refine_od"] = bool(self.refine_od_var.get())
            _isz = self.vars["imgsz"].get().strip()
            self.cfg["imgsz"] = _isz if _isz.lower() == "auto" else int(_isz)
            self.cfg["frames_avg"] = max(1, int(self.vars["frames_avg"].get()))
            self.cfg["tcp_enabled"] = bool(self.tcp_enabled_var.get())
            self.cfg["tcp_host"] = self.vars["tcp_host"].get() or "0.0.0.0"
            self.cfg["tcp_port"] = int(self.vars["tcp_port"].get())
            self.cfg["tcp_format"] = self.vars["tcp_format"].get() or "{id},{x},{y}"
            self.cfg["tcp_empty_message"] = self.vars["tcp_empty_message"].get()
            self.cfg["sort_by"] = self.sort_by_var.get() or "y"
            self.cfg["sort_desc"] = bool(self.sort_desc_var.get())
            self.cfg["robot_ip"] = self.vars["robot_ip"].get() or "127.0.0.1"
            self.cfg["robot_port"] = int(self.vars["robot_port"].get())
            self.cfg["robot_query"] = self.vars["robot_query"].get()
            self.cfg["robot_poll"] = float(self.vars["robot_poll"].get())
            self.cfg["auto_start"] = bool(self.auto_start_var.get())
        except ValueError as e:
            messagebox.showerror("Invalid value",
                                 "Numbers required for poll/offset/detection "
                                 "params and TCP port.\n%s" % e)
            return False
        return True

    def _refresh_fields(self):
        """Push self.cfg values back into the GUI fields."""
        for key, var in self.vars.items():
            if key in self.cfg:
                var.set(str(self.cfg[key]))
        self.tcp_enabled_var.set(bool(self.cfg.get("tcp_enabled")))
        self.sort_by_var.set(self.cfg.get("sort_by", "y"))
        self.sort_desc_var.set(bool(self.cfg.get("sort_desc")))
        self.auto_start_var.set(bool(self.cfg.get("auto_start", True)))
        self.undistort_var.set(bool(self.cfg.get("undistort")))
        self.transform_var.set(self.cfg.get("calib_transform", "homography"))
        self.measure_inner_var.set(bool(self.cfg.get("measure_inner")))
        self.model_type_var.set(self.cfg.get("model_type", "auto"))
        self.subpixel_var.set(bool(self.cfg.get("subpixel", True)))
        self.clahe_var.set(bool(self.cfg.get("clahe", False)))
        self.auto_retry_var.set(bool(self.cfg.get("auto_retry", True)))
        self.multiscale_var.set(bool(self.cfg.get("multiscale", True)))
        self.refine_od_var.set(bool(self.cfg.get("refine_od", True)))

    def save(self, announce=True):
        if not self._read_fields():
            return False
        save_config(self.cfg)
        self._refresh_map_info()
        self._apply_tcp()
        if announce:
            messagebox.showinfo("Saved", "Settings saved to\n%s" % CONFIG_FILE)
        return True

    def export_config(self):
        if not self._read_fields():
            return
        f = filedialog.asksaveasfilename(
            initialdir=ROOT, defaultextension=".json",
            initialfile="ring_config.json",
            filetypes=[("JSON", "*.json"), ("all", "*.*")])
        if not f:
            return
        try:
            json.dump(self.cfg, open(f, "w"), indent=2)
            messagebox.showinfo("Exported", "Configuration exported to\n%s" % f)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))

    def import_config(self):
        f = filedialog.askopenfilename(
            initialdir=ROOT, filetypes=[("JSON", "*.json"), ("all", "*.*")])
        if not f:
            return
        try:
            data = json.load(open(f))
            if not isinstance(data, dict):
                raise ValueError("not a config object")
        except Exception as e:
            messagebox.showerror("Import failed", "Cannot read %s\n%s" % (f, e))
            return
        # keep only known keys, so a stray file can't inject junk
        for k, v in data.items():
            if k in DEFAULT_CONFIG:
                self.cfg[k] = v
        self._refresh_fields()
        save_config(self.cfg)
        self._refresh_map_info()
        self._apply_tcp()
        messagebox.showinfo("Imported",
                            "Configuration imported from\n%s\n\n"
                            "Use Save & Restart if you changed the model." % f)

    def save_and_restart(self):
        if self.save(announce=False):
            self.restart_app(confirm=False)

    def restart_app(self, confirm=True):
        if confirm and not messagebox.askyesno(
                "Restart", "Restart the application now?"):
            return
        try:
            self.worker.stop_flag.set()
            self.robot_reader.stop_flag.set()
            self.tcp.stop()
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass
        # re-launch this same program (fresh process -> model reloads, etc.)
        os.execl(sys.executable, sys.executable, *sys.argv)

    def _apply_tcp(self):
        if self.cfg.get("tcp_enabled"):
            ok = self.tcp.start(self.cfg.get("tcp_host", "0.0.0.0"),
                                self.cfg.get("tcp_port", 5000))
            self.tcp_info.config(
                text="listening on %s:%s" % (self.cfg["tcp_host"],
                                             self.cfg["tcp_port"]) if ok
                else "failed to start (port in use?)",
                foreground="#1a7f37" if ok else "#cf222e")
        else:
            self.tcp.stop()
            self.tcp_info.config(text="disabled", foreground="#555")

    # ---- live actions ----
    def start(self):
        self.worker.start_watching()
        self.status.config(text="Watching", foreground="#1a7f37")
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

    def stop(self):
        self.worker.stop_watching()
        self.status.config(text="Idle", foreground="#333")
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    def open_image(self):
        f = filedialog.askopenfilename(
            initialdir=self.cfg.get("watch_folder", ROOT),
            filetypes=[("images", "*.bmp *.png *.jpg *.jpeg *.tif *.tiff")])
        if f:
            self.worker.submit(f, "process")

    # ---- calibration actions ----
    def calib_grab(self):
        folder = self.cfg.get("watch_folder", "")
        ext = self.cfg.get("image_ext", ".bmp")
        pattern = (self.cfg.get("name_pattern") or "*").lower()
        files = [f for f in glob.glob(os.path.join(folder, "*" + ext))
                 if not f.lower().endswith("_rings.png")
                 and fnmatch.fnmatch(os.path.basename(f).lower(), pattern)]
        if not files:
            messagebox.showwarning("No image",
                                   "No %s image found in\n%s" % (ext, folder))
            return
        newest = max(files, key=os.path.getmtime)
        self.worker.submit(newest, "calib")

    def calib_batch(self):
        self._read_fields()          # so the map path / model / undistort are current
        self.cfg["batch_multi"] = self.batch_multi_var.get() or "largest"
        folder = filedialog.askdirectory(
            title="Folder with calibration images + Excel/CSV",
            initialdir=self.cfg.get("watch_folder", ROOT))
        if not folder:
            return
        self._log("batch calibration from %s ..." % folder)
        self.worker.submit(folder, "batchcal")

    def calib_load(self):
        f = filedialog.askopenfilename(
            initialdir=self.cfg.get("watch_folder", ROOT),
            filetypes=[("images", "*.bmp *.png *.jpg *.jpeg *.tif *.tiff")])
        if f:
            self.worker.submit(f, "calib")

    def _selected_ring(self):
        if not self._calib_rings:
            return None
        try:
            idx = int(self.calib_ring_sel.get()) - 1
        except ValueError:
            idx = 0
        idx = max(0, min(idx, len(self._calib_rings) - 1))
        return self._calib_rings[idx]

    def download_model(self):
        name = self.vars["model"].get() or "FastSAM-x.pt"
        self._log("downloading model '%s' (needs internet) ..." % name)

        def work():
            try:
                p = download_asset(name, ROOT, lambda m: self.q.put(("log", m)))
                self.q.put(("log", "model ready: %s" % p))
            except Exception as e:
                self.q.put(("log", "model download failed: %s" % e))
        threading.Thread(target=work, daemon=True).start()

    def _toggle_robot_read(self):
        # make sure the reader uses the latest IP/port/command from the fields
        self._read_fields()
        if self.robot_live_var.get():
            self.robot_reader.enabled.set()
            self._log("robot position read: ON (%s:%s)"
                      % (self.cfg.get("robot_ip"), self.cfg.get("robot_port")))
        else:
            self.robot_reader.enabled.clear()

    # ---- calibration profiles ----
    def _profiles_dir(self):
        d = os.path.join(ROOT, "config", "profiles")
        os.makedirs(d, exist_ok=True)
        return d

    @staticmethod
    def _is_map_file(path):
        """True if the JSON looks like a calibration map (has H or M)."""
        try:
            d = json.load(open(path))
            return isinstance(d, dict) and ("H" in d or "M" in d)
        except Exception:
            return False

    def _refresh_profiles(self):
        self._profile_paths = {}
        names = []

        def add(path):
            if not path or not os.path.exists(path):
                return
            n = os.path.basename(path)
            if n not in self._profile_paths and self._is_map_file(path):
                self._profile_paths[n] = path
                names.append(n)

        # every map in config/ and config/profiles/, plus the active file
        add(self.vars["homography_file"].get())
        add(os.path.join(ROOT, "config", "robot_map.json"))
        for folder in (os.path.join(ROOT, "config"), self._profiles_dir()):
            for f in sorted(glob.glob(os.path.join(folder, "*.json"))):
                add(f)

        self.profile_combo["values"] = names
        cur = os.path.basename(self.vars["homography_file"].get() or "")
        if cur in names:
            self.profile_var.set(cur)

    def _profile_selected(self):
        name = self.profile_var.get()
        path = self._profile_paths.get(name)
        if not path:
            return
        self.vars["homography_file"].set(path)
        self.cfg["homography_file"] = path
        save_config(self.cfg)
        self._refresh_map_info()
        self._log("active calibration profile: %s" % name)

    def calib_save_as_profile(self):
        name = simpledialog.askstring("Save profile",
                                      "Profile name (e.g. cam1_fixtureA):")
        if not name:
            return
        name = name if name.lower().endswith(".json") else name + ".json"
        path = os.path.join(self._profiles_dir(), name)
        self.vars["homography_file"].set(path)   # calib_fit saves to this path
        self.calib_fit()
        self._refresh_profiles()

    def calib_verify(self):
        ring = self._selected_ring()
        if ring is None:
            self.verify_lbl.config(text="grab/load an image with a ring first",
                                   foreground="#b26b00")
            return
        mapper = load_mapper(self.vars["homography_file"].get())
        if mapper is None:
            self.verify_lbl.config(text="no valid homography loaded",
                                   foreground="#cf222e")
            return
        px, py, _ = ring
        prx, pry = map_point(mapper, px, py)     # predicted robot XY (no offset)
        live = self.robot_reader.get_latest()
        if live is None:
            self.verify_lbl.config(
                text="predicted (%.2f, %.2f) mm  |  no live robot position "
                     "(enable 'Read robot position')" % (prx, pry),
                foreground="#b26b00")
            return
        err = ((prx - live[0]) ** 2 + (pry - live[1]) ** 2) ** 0.5
        self.verify_lbl.config(
            text="predicted (%.2f, %.2f)  |  live (%.2f, %.2f)  |  error %.2f mm"
                 % (prx, pry, live[0], live[1], err),
            foreground="#1a7f37" if err < 3 else "#cf222e")

    def calib_update_from_live(self):
        latest = self.robot_reader.get_latest()
        if latest is None:
            self._log("no live robot position yet "
                      "(enable 'Read robot position (TCP)' and wait for a value)")
            return
        x, y = latest
        self.rx_var.set("%.3f" % x)
        self.ry_var.set("%.3f" % y)

    def calib_add(self):
        ring = self._selected_ring()
        if ring is None:
            messagebox.showwarning("No ring", "Grab/Load an image with a ring first.")
            return
        try:
            rx = float(self.rx_var.get())
            ry = float(self.ry_var.get())
        except ValueError:
            messagebox.showerror("Invalid", "Robot X and Y must be numbers.")
            return
        px, py, _ = ring
        self.points.append((px, py, rx, ry))
        self._refresh_points()
        self.rx_var.set("")
        self.ry_var.set("")

    def calib_remove(self):
        sel = self.ptree.selection()
        if not sel:
            return
        idx = self.ptree.index(sel[0])
        del self.points[idx]
        self._refresh_points()

    def calib_clear(self):
        self.points = []
        self._refresh_points()

    def _refresh_points(self):
        for row in self.ptree.get_children():
            self.ptree.delete(row)
        for i, (px, py, rx, ry) in enumerate(self.points):
            self.ptree.insert("", tk.END, values=(i + 1, px, py, rx, ry))
        n = len(self.points)
        self.calib_count.config(
            text="Points: %d  (need >= 4)" % n,
            foreground="#1a7f37" if n >= 4 else "#b26b00")

    def calib_load_points(self):
        """Load the points from the current map so it can be extended/improved."""
        path = self.vars["homography_file"].get()
        try:
            pts = json.load(open(path)).get("points", [])
        except Exception as e:
            messagebox.showerror("Load failed", "Cannot read %s\n%s" % (path, e))
            return
        if not pts:
            messagebox.showwarning("No points", "That map file has no saved points.")
            return
        self.points = [(p["px"], p["py"], p["robot_x"], p["robot_y"])
                       for p in pts]
        self._refresh_points()
        self._log("loaded %d points from %s (add more, then Fit & Save)"
                  % (len(self.points), os.path.basename(path)))

    def calib_coverage(self):
        """Show where the current map is accurate vs weak, to guide improvement."""
        path = self.vars["homography_file"].get()
        try:
            d = json.load(open(path))
        except Exception as e:
            messagebox.showerror("Coverage", "Cannot read %s\n%s" % (path, e))
            return
        pts = d.get("points", [])
        mapper = load_mapper(path)
        if not pts or mapper is None:
            messagebox.showwarning("Coverage", "Map has no points to assess.")
            return
        P = [(p["px"], p["py"]) for p in pts]
        R = [(p["robot_x"], p["robot_y"]) for p in pts]
        err = []
        for (px, py), (rx, ry) in zip(P, R):
            mx, my = map_point(mapper, px, py)
            err.append((mx - rx) ** 2 + (my - ry) ** 2)
        err = [e ** 0.5 for e in err]
        wh = d.get("frame_wh") or getattr(self, "_calib_wh", None)
        if wh:
            W, H = int(wh[0]), int(wh[1])
        else:
            W = int(max(p[0] for p in P) + 40)
            H = int(max(p[1] for p in P) + 40)
        # coverage heat: green where near a calibration point, red where far
        mask = np.full((H, W), 255, np.uint8)
        for px, py in P:
            xi, yi = int(np.clip(px, 0, W - 1)), int(np.clip(py, 0, H - 1))
            mask[yi, xi] = 0
        dist = cv2.distanceTransform(mask, cv2.DIST_L2, 5)
        t = np.clip(dist / (0.20 * np.hypot(W, H)), 0, 1)      # 0 near .. 1 far
        heat = np.zeros((H, W, 3), np.uint8)
        heat[..., 1] = (150 * (1 - t)).astype(np.uint8)        # green = covered
        heat[..., 2] = (200 * t).astype(np.uint8)              # red = uncovered
        # overlay the heat on the last captured calibration image, if wanted
        base = getattr(self, "_calib_last_bgr", None)
        if self.cov_overlay_var.get() and base is not None \
                and base.shape[:2] == (H, W):
            img = cv2.addWeighted(base, 0.55, heat, 0.55, 0)
        else:
            img = heat
        for (px, py), e in zip(P, err):
            col = (0, 230, 0) if e < 1 else ((0, 180, 255) if e < 2 else (0, 0, 255))
            cv2.circle(img, (int(px), int(py)), 6, col, -1)
            cv2.circle(img, (int(px), int(py)), 6, (255, 255, 255), 1)
        cv2.putText(img, "green=covered  red=add points here  dots: <1 /1-2 />2mm",
                    (6, 16), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (255, 255, 255), 1,
                    cv2.LINE_AA)
        loo = d.get("loo_rms_mm")
        cv2.putText(img, "model %s  RMS %.2f  LOO %s mm"
                    % (d.get("transform", "?"), d.get("rms_mm", 0),
                       ("%.2f" % loo) if loo else "n/a"),
                    (6, H - 8), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (255, 255, 255), 1,
                    cv2.LINE_AA)
        out = os.path.splitext(path)[0] + "_coverage.png"
        try:
            cv2.imwrite(out, img)
        except Exception:
            pass
        try:
            self._calib_imgtk = self._to_tk(img, (560, 470))
            self.calib_canvas.config(image=self._calib_imgtk)
        except Exception as e:
            self._log("coverage display error: %s" % e)
        self._log("coverage map shown (saved %s). Red areas need more points."
                  % os.path.basename(out))

    def calib_fit(self):
        # distortion-aware if enabled and intrinsics available
        K = dist = None
        note = ""
        if self.undistort_var.get():
            K, dist = load_intrinsics(self.vars["intrinsics_file"].get())
            if K is None:
                messagebox.showwarning(
                    "No intrinsics",
                    "Distortion-aware is on but the intrinsics file is missing/"
                    "invalid.\nFitting without undistortion.")
            else:
                note = "  (distortion-aware)"
        kind = self.transform_var.get() or "homography"
        try:
            result = fit_transform(self.points, kind, K=K, dist=dist,
                                   frame_wh=getattr(self, "_calib_wh", None))
        except ValueError as e:
            messagebox.showerror("Cannot fit", str(e))
            return
        path = self.vars["homography_file"].get() or self.cfg["homography_file"]
        try:
            os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
            json.dump(result, open(path, "w"), indent=2)
        except Exception as e:
            messagebox.showerror("Save failed", str(e))
            return
        loo = ("%.2f" % result["loo_rms_mm"]) if result["loo_rms_mm"] else "n/a"
        msg = ("Saved %s%s\npoints used: %d   fit RMS: %.2f mm   LOO: %s mm"
               % (os.path.basename(path), note, result["points_used"],
                  result["rms_mm"], loo))
        msg += "\nmodel: %s" % kind
        if "angle_deg" in result:
            msg += "  (angle %.2f deg, scale %.4f mm/px)" % (
                result["angle_deg"], result["scale_mm_px"])
        if result["excluded"]:
            msg += "\nexcluded (check robot XY): %s" % result["excluded"]
        if result.get("coverage_warning"):
            msg += "\n[!] " + result["coverage_warning"]
        self.calib_result.config(text=msg, foreground="#1a7f37")
        self._refresh_map_info()
        self._refresh_profiles()
        self.cfg["calib_transform"] = kind
        save_config(self.cfg)
        messagebox.showinfo("Calibration saved", msg)

    def on_close(self):
        self.worker.stop_flag.set()
        self.robot_reader.stop_flag.set()
        self.tcp.stop()
        self.root.after(100, self.root.destroy)

    # ---- pump ----
    def _pump(self):
        try:
            while True:
                kind, payload = self.q.get_nowait()
                if kind == "log":
                    self._log(payload)
                elif kind == "ready":
                    self.status.config(text="Idle (model ready)",
                                       foreground="#333")
                    self.start_btn.config(state=tk.NORMAL)
                    if self.cfg.get("auto_start", True):
                        self.start()   # begin watching automatically on open
                elif kind == "result":
                    self._show(payload)
                elif kind == "calib_result":
                    self._show_calib(payload)
                elif kind == "batchcal_result":
                    self._show_batchcal(payload)
                elif kind == "robot_live":
                    x, y = payload
                    self.robot_live_lbl.config(
                        text="live: X=%.3f  Y=%.3f" % (x, y), foreground="#1a7f37")
                elif kind == "robot_status":
                    if payload not in ("off",):
                        self._log("robot: %s" % payload)
                    if payload == "off" or payload.startswith("disconnected"):
                        self.robot_live_lbl.config(text="live: -- (%s)" % payload,
                                                   foreground="#777")
        except queue.Empty:
            pass
        self._update_tcp_status()
        self.root.after(150, self._pump)

    def _update_tcp_status(self):
        running, addrs = self.tcp.info()
        if not running:
            self.tcp_status.config(text="TCP: off", foreground="#777")
        elif not addrs:
            self.tcp_status.config(text="TCP: listening, 0 clients",
                                   foreground="#b26b00")
        else:
            shown = ", ".join(addrs[:2]) + ("..." if len(addrs) > 2 else "")
            self.tcp_status.config(
                text="TCP: %d client(s) [%s]" % (len(addrs), shown),
                foreground="#1a7f37")

    def _sort_tree(self, col):
        desc = not self._tree_sort[1] if self._tree_sort[0] == col else False
        self._tree_sort = (col, desc)
        rows = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]

        def key(v):
            try:
                return float(v[0])
            except ValueError:
                return v[0]
        rows.sort(key=key, reverse=desc)
        for i, (_, k) in enumerate(rows):
            self.tree.move(k, "", i)

    def _log(self, m):
        self.logbox.config(state=tk.NORMAL)
        self.logbox.insert(tk.END, time.strftime("[%H:%M:%S] ") + m + "\n")
        self.logbox.see(tk.END)
        self.logbox.config(state=tk.DISABLED)

    def _to_tk(self, bgr, box):
        rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        im = Image.fromarray(rgb)
        im.thumbnail(box)
        return ImageTk.PhotoImage(im)

    def _zoom(self, delta):
        self.zoom = max(0.25, min(8.0, self.zoom + delta))
        self._render_live()

    def _zoom_set(self, z):
        self.zoom = z
        self._render_live()

    def _zoom_fit(self):
        if self._last_bgr is None:
            return
        h, w = self._last_bgr.shape[:2]
        cw = self.img_canvas.winfo_width() or 560
        ch = self.img_canvas.winfo_height() or 460
        self.zoom = max(0.1, min(cw / w, ch / h))
        self._render_live()

    def _render_live(self):
        if self._last_bgr is None:
            return
        h, w = self._last_bgr.shape[:2]
        nw, nh = max(1, int(w * self.zoom)), max(1, int(h * self.zoom))
        try:
            rgb = cv2.cvtColor(self._last_bgr, cv2.COLOR_BGR2RGB)
            im = Image.fromarray(rgb).resize((nw, nh))
            self._imgtk = ImageTk.PhotoImage(im)
            self.img_canvas.itemconfig(self._img_item, image=self._imgtk)
            self.img_canvas.configure(scrollregion=(0, 0, nw, nh))
            self.zoom_lbl.config(text="%d%%" % int(self.zoom * 100))
        except Exception as e:
            self._log("display error: %s" % e)

    def _show(self, res):
        if res.get("empty"):
            self.banner.config(text="CONVEYOR EMPTY  -  %s" % res["name"],
                               bg="#cf222e")
        else:
            self.banner.config(text="%d ring(s)  -  %s"
                                    % (len(res["rings"]), res["name"]),
                               bg="#1a7f37")
        if "total_ms" in res:
            self.timing_lbl.config(
                text="detect %.0f ms | total %.0f ms"
                     % (res.get("detect_ms", 0), res["total_ms"]))
        self._last_bgr = res["image"]
        self._render_live()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for r in res["rings"]:
            self.tree.insert("", tk.END, values=(
                r["id"], r["x"], r["y"], r["diameter"],
                r["robot_x"], r["robot_y"], r.get("diameter_mm", ""),
                r.get("inner_dia", ""), r.get("inner_dia_mm", "")))

    def _show_batchcal(self, res):
        if res.get("error"):
            miss = res.get("missing")
            m = "Batch calibration failed:\n%s" % res["error"]
            if miss:
                m += "\nskipped: %s" % ", ".join(miss[:8])
            messagebox.showerror("Batch calibration", m)
            self.calib_result.config(text=m, foreground="#cf222e")
            return
        loo = ("%.2f" % res["loo_rms_mm"]) if res.get("loo_rms_mm") else "n/a"
        msg = ("Batch calibration saved: %s\ntable: %s   points used: %d   "
               "model: %s\nfit RMS: %.2f mm   LOO: %s mm"
               % (os.path.basename(res["_saved_to"]), res.get("_table", "?"),
                  res["points_used"], res.get("transform", "?"),
                  res["rms_mm"], loo))
        if res.get("angle_deg") is not None:
            msg += "  (angle %.2f deg, scale %.4f mm/px)" % (
                res["angle_deg"], res["scale_mm_px"])
        if res.get("excluded"):
            msg += "\nexcluded (check robot XY): %s" % res["excluded"]
        if res.get("_multi"):
            msg += "\n[!] %d image(s) had MULTIPLE rings (used '%s'): %s" % (
                len(res["_multi"]), res.get("_multi_strategy", "largest"),
                ", ".join(res["_multi"][:6]))
        if res.get("_missing"):
            msg += "\nskipped %d image(s): %s" % (
                len(res["_missing"]), ", ".join(res["_missing"][:6]))
        if res.get("coverage_warning"):
            msg += "\n[!] " + res["coverage_warning"]
        self.calib_result.config(text=msg, foreground="#1a7f37")
        self._refresh_map_info()
        self._refresh_profiles()
        messagebox.showinfo("Batch calibration", msg)

    def _show_calib(self, res):
        self._calib_rings = res["rings"]
        h, w = res["image"].shape[:2]
        self._calib_wh = (w, h)
        self._calib_last_bgr = res["image"].copy()   # for coverage overlay
        try:
            self._calib_imgtk = self._to_tk(res["image"], (520, 430))
            self.calib_canvas.config(image=self._calib_imgtk)
        except Exception as e:
            self._log("calib display error: %s" % e)
        n = len(self._calib_rings)
        self.calib_ring_sel.config(from_=1, to=max(1, n))
        if n:
            px, py, _ = self._calib_rings[0]
            self.calib_pix.config(text="pixel: (%.1f, %.1f)  [%d ring(s)]"
                                       % (px, py, n))
        else:
            self.calib_pix.config(text="pixel: -  (no ring found)")


def main():
    global tk, ttk, filedialog, messagebox, simpledialog, Image, ImageTk
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, simpledialog
    from PIL import Image, ImageTk
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
